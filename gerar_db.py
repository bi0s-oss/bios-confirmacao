"""
Regenera db.json (usado pelo site) a partir dos HTMLs em html_cvs/ e das abas
de Piloto_att.xlsx.

Faz 4 coisas:
  1. Remove pesquisadores que saíram do BI0S (EXCLUIR_IDS).
  2. Extrai produção/orientações/bancas/etc. dos HTMLs, já nos nomes de campo
     que index.html (renderProducao) espera.
  3. Atualiza "projetos" a partir da aba Projetos de Piloto_att.xlsx.
  4. Acrescenta às categorias de produção os itens mais recentes cadastrados
     manualmente nas abas de produção de Piloto_att.xlsx (Artigos, Livros,
     Capítulos, Trabalhos em anais, Patentes, Software, Produção técnica,
     Bancas, Orientações, Eventos, Prêmios) - essas abas têm itens de
     2025/2026 que ainda não estão nos HTMLs salvos.
"""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup

import gerar_tabela as gt

HTML_DIR = Path("html_cvs")
OUTPUT_FILE = Path("db.json")
PILOTO_XLSX = Path("Piloto_att.xlsx")

EXCLUIR_IDS = {
    "1827990606590127",  # Breno Bernard Nicolau de Franca - saiu do BI0S em 24/02/26
    "7917967295254962",  # Igor Gadelha Pereira - desligado, retirar
    "1610878342077626",  # Marcos Ricardo Omena de Albuquerque Maximo - saiu em 01/12/2023
    "5803743950533597",  # Maria Leticia Cintra - pediu para ser retirada do BI0S
}


def textos(objs):
    return [o["texto"] for o in objs if o]


def dedup(lst):
    seen, out = set(), []
    for t in lst:
        k = t.lower()[:100]
        if k and k not in seen:
            seen.add(k)
            out.append(t)
    return out


def extrair_orientacoes_split(dc):
    """Orientações concluídas e em andamento vivem no MESMO data-cell,
    separadas por dois headers inst_back distintos."""
    children = [c for c in dc.children if hasattr(c, "name") and c.name]
    concluidas, andamento = [], []
    bucket = None
    for child in children:
        cls = set(child.get("class") or [])
        txt = child.get_text()
        if "inst_back" in cls:
            t = gt.limpar(txt).lower()
            if "andamento" in t:
                bucket = "andamento"
            elif "conclu" in t:
                bucket = "concluidas"
            else:
                bucket = None
        elif "layout-cell-11" in cls and bucket:
            t = gt.limpar(txt)
            if t:
                (andamento if bucket == "andamento" else concluidas).append(t)
    return concluidas, andamento


def extrair_cv_db(html_path):
    html = html_path.read_text(encoding="utf-8")
    soup = BeautifulSoup(html, "html.parser")
    data_cells = soup.find_all("div", class_="data-cell")

    artigos, livros, capitulos, anais, apresentacoes, producao_tecnica = [], [], [], [], [], []
    patentes, software, bancas, eventos, premios = [], [], [], [], []
    orient_concluidas, orient_andamento = [], []

    for dc in data_cells:
        txt = dc.get_text()[:300]

        if ("bibliogr" in txt.lower() or "Produ" in txt[:60]) and ("Patente" not in txt[:60]):
            if gt.cells_direct_count(dc) > 0 or len(dc.find_all("div", class_="layout-cell-11")) >= 10:
                secoes = gt.extrair_producao_bibliografica(dc)
                artigos.extend(textos(secoes.get("artigos", [])))
                livros.extend(textos(secoes.get("livros", [])))
                capitulos.extend(textos(secoes.get("capitulos", [])))
                anais.extend(textos(secoes.get("anais", [])))
                anais.extend(textos(secoes.get("resumos_anais", [])))
                apresentacoes.extend(textos(secoes.get("apresentacoes", [])))
                producao_tecnica.extend(textos(secoes.get("producao_tecnica", [])))

        elif ("Patente" in txt or "Programa de computador" in txt or "Programa de Computador" in txt) and gt.cells_direct_count(dc) > 0:
            pats, softs = gt.extrair_patentes_e_software(dc)
            patentes.extend(textos(pats))
            software.extend(textos(softs))

        elif "bancas" in txt.lower() and "trabalhos" in txt.lower():
            bancas.extend(textos(gt.extrair_items_dc(dc)))

        elif "Orienta" in txt and ("andamento" in txt.lower() or "conclu" in txt.lower()):
            c, a = extrair_orientacoes_split(dc)
            orient_concluidas.extend(c)
            orient_andamento.extend(a)

        elif "eventos" in txt.lower() and "congressos" in txt.lower():
            eventos.extend(textos(gt.extrair_items_dc(dc)))

        elif ("mio" in txt.lower() or "tulo" in txt.lower()) and "Premio" in txt or "Pr" in txt and "mio" in txt:
            premios.extend(textos(gt.extrair_premios(dc)))

    return {
        "artigosCompletos": dedup(artigos),
        "trabalhosCongresso": dedup(anais),
        "livrosCapitulos": dedup(livros + capitulos),
        "patentesRegistros": dedup(patentes),
        "softwareRegistros": dedup(software),
        "orientacoesConcluidas": dedup(orient_concluidas),
        "orientacoesAndamento": dedup(orient_andamento),
        "premiosTitulos": dedup(premios),
        "eventos": dedup(eventos),
        "bancas": dedup(bancas),
        "producaoTecnica": dedup(producao_tecnica),
        "apresentacoes": dedup(apresentacoes),
    }


def normalizar_nome(s):
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z\s]", "", s)
    return re.sub(r"\s+", " ", s).strip().lower()


NOME_PARA_ID = {normalizar_nome(nome): id_ for id_, nome in gt.PESQUISADORES.items()}

# aba do Piloto -> campo de db.json (Livros e Capítulos caem no mesmo campo)
SHEET_PARA_CAMPO = {
    "Artigos em periódicos": "artigosCompletos",
    "Livros publicados": "livrosCapitulos",
    "Capítulos de livros": "livrosCapitulos",
    "Trabalhos em anais": "trabalhosCongresso",
    "Patentes e registros": "patentesRegistros",
    "Software registrado": "softwareRegistros",
    "Produção técnica": "producaoTecnica",
    "Participação em bancas": "bancas",
    "Eventos": "eventos",
}


def carregar_producao_piloto():
    """Le as abas de producao de Piloto_att.xlsx e devolve {id: {campo: [textos]}}."""
    wb = openpyxl.load_workbook(PILOTO_XLSX, read_only=True)
    extra = {}

    def add(id_, campo, texto):
        extra.setdefault(id_, {}).setdefault(campo, []).append(texto)

    for sheet, campo in SHEET_PARA_CAMPO.items():
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[2]:
                continue
            id_ = NOME_PARA_ID.get(normalizar_nome(str(row[0])))
            if not id_:
                continue
            add(id_, campo, gt.limpar(str(row[2])))

    # Premios usam o formato "ano — descricao", como o resto do codigo
    ws = wb["Prêmios e títulos"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[2]:
            continue
        id_ = NOME_PARA_ID.get(normalizar_nome(str(row[0])))
        if not id_:
            continue
        ano, desc = row[1], gt.limpar(str(row[2]))
        texto = f"{ano} — {desc}" if ano else desc
        add(id_, "premiosTitulos", texto)

    # Orientacoes: "Inicio" no texto = em andamento, senao = concluida
    ws = wb["Orientações"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[2]:
            continue
        id_ = NOME_PARA_ID.get(normalizar_nome(str(row[0])))
        if not id_:
            continue
        desc = gt.limpar(str(row[2]))
        campo = "orientacoesAndamento" if "início" in desc.lower() or "inicio" in desc.lower() else "orientacoesConcluidas"
        add(id_, campo, desc)

    return extra


def carregar_projetos():
    wb = openpyxl.load_workbook(PILOTO_XLSX, read_only=True)
    ws = wb["Projetos"]
    projetos_por_id = {}
    ultima_at_por_id = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        id_, _nome, ultima_at, periodo, titulo = row[0], row[1], row[2], row[3], row[4]
        id_ = str(id_).strip()
        if not titulo:
            continue
        projetos_por_id.setdefault(id_, []).append({"periodo": periodo or "", "titulo": titulo})
        if ultima_at:
            ultima_at_por_id[id_] = str(ultima_at)
    return projetos_por_id, ultima_at_por_id


def dedup_projetos(lista):
    seen, out = set(), []
    for p in lista:
        k = (p["periodo"], p["titulo"].lower()[:100])
        if k not in seen:
            seen.add(k)
            out.append(p)
    return out


def carregar_projetos_atuais():
    """Projetos já presentes no db.json atual, usados como fallback para
    pesquisadores que não aparecem na aba Projetos de Piloto_att.xlsx."""
    if not OUTPUT_FILE.exists():
        return {}
    with open(OUTPUT_FILE, encoding="utf-8") as f:
        atual = json.load(f)
    return {p["id"]: p.get("projetos", []) for p in atual}


def main():
    projetos_por_id, ultima_at_por_id = carregar_projetos()
    projetos_atuais_por_id = carregar_projetos_atuais()
    producao_piloto_por_id = carregar_producao_piloto()
    htmls = sorted(HTML_DIR.glob("*.html"))

    resultado = []
    excluidos = []
    sem_projeto_na_planilha = []
    adicionados_piloto = {}

    for html_path in htmls:
        id_lattes = html_path.stem
        if id_lattes in EXCLUIR_IDS:
            excluidos.append(id_lattes)
            continue

        nome = gt.PESQUISADORES.get(id_lattes, "ID_" + id_lattes)
        print("  " + nome)

        cats = extrair_cv_db(html_path)

        extra = producao_piloto_por_id.get(id_lattes, {})
        for campo, novos_itens in extra.items():
            antes = len(cats.get(campo, []))
            cats[campo] = dedup(cats.get(campo, []) + novos_itens)
            adicionados_piloto[campo] = adicionados_piloto.get(campo, 0) + (len(cats[campo]) - antes)

        if id_lattes in projetos_por_id:
            projetos = dedup_projetos(projetos_por_id[id_lattes])
        else:
            sem_projeto_na_planilha.append(nome)
            projetos = projetos_atuais_por_id.get(id_lattes, [])

        ultima_at = ultima_at_por_id.get(id_lattes, "")
        if not ultima_at:
            html = html_path.read_text(encoding="utf-8")
            soup = BeautifulSoup(html, "html.parser")
            ultima_at = gt.extrair_ultima_atualizacao(soup)

        entry = {
            "id": id_lattes,
            "nome": nome,
            "ultima_atualizacao": ultima_at,
            "projetos": projetos,
        }
        entry.update(cats)
        resultado.append(entry)

    resultado.sort(key=lambda p: p["nome"])

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, separators=(",", ":"))

    print("\nExcluidos (%d): %s" % (len(excluidos), excluidos))
    print("Salvos: %d pesquisadores em %s" % (len(resultado), OUTPUT_FILE))
    print("\nItens novos incorporados do Piloto_att.xlsx (apos dedup):")
    for campo, qtd in sorted(adicionados_piloto.items()):
        print("  %s: +%d" % (campo, qtd))
    if sem_projeto_na_planilha:
        print("\nSem projetos na aba 'Projetos' do Piloto_att.xlsx (%d) - ficaram com projetos=[]:" % len(sem_projeto_na_planilha))
        for n in sem_projeto_na_planilha:
            print("  - " + n)


if __name__ == "__main__":
    main()
